import os
import openpyxl

# 设置输入和输出文件夹路径
input_folder = "E:/github/new-productivity/criterion/excel"
output_file = "E:/github/new-productivity/criterion/file.xlsx"

# 创建一个新的 workbook
output_workbook = openpyxl.Workbook()
output_sheet = output_workbook.active

# 遍历输入文件夹中的所有 Excel 文件
for filename in os.listdir(input_folder):
    if filename.endswith(".xlsx"):
        # 打开当前 Excel 文件
        input_file = os.path.join(input_folder, filename)
        input_workbook = openpyxl.load_workbook(input_file)
        input_sheet = input_workbook.active

        # 读取当前 Excel 文件中的所有行并添加到输出 Excel 文件中
        for row in range(1, input_sheet.max_row + 1):
            row_data = [cell.value for cell in input_sheet[row]]
            row_data.append(filename)  # 添加一列表示数据来自哪个 Excel 文件
            output_sheet.append(row_data)

# 保存输出 Excel 文件
output_workbook.save(output_file)
print("Excel files merged successfully!")
